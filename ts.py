# -----------------------------------------------------------------------------
# Time Sheet Tracker
#
# A simple Python GUI app to clock in/out and save session notes to Excel.
# The application uses tkinter for UI and openpyxl for Excel handling.
#
# Author: Rong Zheng
# Date: 2025-06-30
#
# License: Creative Commons Attribution 4.0 International (CC BY 4.0)
# You are free to share and adapt this work, even commercially, as long as
# appropriate credit is given.
# https://creativecommons.org/licenses/by/4.0/
#
# Acknowledgment: This code was developed in part with the assistance of
# ChatGPT by OpenAI.
# -----------------------------------------------------------------------------

import tkinter as tk
from tkinter import messagebox
from datetime import datetime
import os
import openpyxl

EXCEL_FILE = 'timesheet.xlsx'

def create_excel_if_needed():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Timesheet'
        ws.append(['Date', 'Clock In', 'Clock Out', 'Note'])
        wb.save(EXCEL_FILE)

def check_active_session():
    if not os.path.exists(EXCEL_FILE):
        return False

    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    now = datetime.now()
    today = now.date().isoformat()

    for row in reversed(list(ws.iter_rows(min_row=2, values_only=False))):
        row_date = row[0].value

        clock_out = row[2].value
        if (row_date == today) and (clock_out is None or str(clock_out).strip() == ''):
            return True  # active session found

    return False  # no active session

def log_entry(clock_in, note=''):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    now = datetime.now()
    today = now.date().isoformat()
    time_str = now.strftime("%H:%M:%S")

    if clock_in:
        ws.append([today, time_str, '', note])
    else:
        # Fill the last open clock-in session for today
        for row in reversed(list(ws.iter_rows(min_row=2, values_only=False))):
            if (row[2].value is None or str(row[2].value).strip() == '') and row[0].value == today:
                row[2].value = time_str
                if note:
                    row[3].value = note

                break
        else:
            messagebox.showwarning("Warning", "No active clock-in found for today.")
            return
    wb.save(EXCEL_FILE)

def toggle_clock():
    global clocked_in
    if clocked_in:
        note = note_box.get("1.0", tk.END).strip()
        log_entry(clock_in=False, note=note)
        status_var.set("Clocked out")
        button.config(text="Clock In")
        note_box.delete("1.0", tk.END)
        note_box.grid_remove()
        note_label.grid_remove()
    else:
        log_entry(clock_in=True)
        status_var.set("Clocked in")
        button.config(text="Clock Out")
        note_label.grid(row=1, column=0, sticky="ne", pady=5)
        note_box.grid(row=1, column=1, pady=5)

    clocked_in = not clocked_in

def on_closing():
    if clocked_in:
        if messagebox.askyesno("Exit", "You're still clocked in. Do you want to clock out before exiting?"):
            toggle_clock()  # Clock out automatically
        else:
            if not messagebox.askokcancel("Exit", "This session will remain open in the timesheet. Exit anyway?"):
                return
    root.destroy()

# Initial state
create_excel_if_needed()
clocked_in = check_active_session()

# GUI
root = tk.Tk()
root.title("Time Sheet Tracker")

frame = tk.Frame(root, padx=10, pady=10)
frame.pack()

button = tk.Button(frame, text="Clock Out" if clocked_in else "Clock In", width=20, command=toggle_clock)
button.grid(row=0, column=0, columnspan=2, pady=10)

note_label = tk.Label(frame, text="Note:")
note_box = tk.Text(frame, width=40, height=5)

status_var = tk.StringVar()
status_var.set("Clocked in" if clocked_in else "Clocked out")
status_label = tk.Label(frame, textvariable=status_var, fg="red")
status_label.grid(row=2, column=0, columnspan=2)

if clocked_in:
    note_label.grid(row=1, column=0, sticky="ne", pady=5)
    note_box.grid(row=1, column=1, pady=5)

# Handle closing
root.protocol("WM_DELETE_WINDOW", on_closing)

root.mainloop()

